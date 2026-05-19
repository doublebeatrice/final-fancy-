import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def ymd(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return ""


def slug(value):
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", str(value).lower())).strip("_")


def ascii_phrases(value):
    text = str(value or "")
    parts = re.split(r"[、,;/；，。()（）:：]+", text)
    phrases = []
    for part in parts:
        cleaned = re.sub(r"[^A-Za-z0-9' -]+", " ", part)
        cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
        if len(cleaned) >= 4 and re.search(r"[a-z]", cleaned):
            phrases.append(cleaned)
    return list(dict.fromkeys(phrases))


def normalize_name_term(name):
    term = re.sub(r"\s*/\s*.*$", "", str(name or ""))
    term = re.sub(r"\s+", " ", term).strip().lower()
    if not term:
        return ""
    if not re.search(r"\bgift|gifts|season|week|day|month\b", term):
        term = f"{term} gifts"
    elif not re.search(r"\bgift|gifts\b", term) and re.search(r"\bday|week|month\b", term):
        term = f"{term} gifts"
    return term


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--out", type=Path, default=Path("data/season_events_2026.json"))
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        if not name:
            continue
        zh_name = ws.cell(row_idx, 2).value or ""
        description = ws.cell(row_idx, 3).value or ""
        date_text = ws.cell(row_idx, 4).value or ""
        date_note = ws.cell(row_idx, 5).value or ""
        product_direction = ws.cell(row_idx, 6).value or ""
        phrases = ascii_phrases(product_direction)
        core_term = phrases[0] if phrases else normalize_name_term(name)
        title_terms = list(dict.fromkeys([
            str(name).strip(),
            core_term,
            *phrases[:8],
        ]))
        event = {
            "key": slug(name),
            "name": str(name).strip(),
            "zhName": str(zh_name).strip(),
            "description": str(description).strip(),
            "dateText": str(date_text).strip(),
            "dateNote": str(date_note).strip(),
            "productDirection": str(product_direction).strip(),
            "coreTerm": core_term,
            "titleTerms": title_terms,
            "nodeStart": ymd(ws.cell(row_idx, 12).value),
            "nodeEnd": ymd(ws.cell(row_idx, 13).value),
            "firstStart": ymd(ws.cell(row_idx, 14).value),
            "firstEnd": ymd(ws.cell(row_idx, 15).value),
            "secondStart": ymd(ws.cell(row_idx, 16).value),
            "secondEnd": ymd(ws.cell(row_idx, 17).value),
            "highFrequencyStart": ymd(ws.cell(row_idx, 18).value),
            "highFrequencyEnd": ymd(ws.cell(row_idx, 19).value),
            "sourceRow": row_idx,
        }
        if event["nodeStart"] and event["nodeEnd"]:
            rows.append(event)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "source": str(args.xlsx),
        "out": str(args.out),
        "events": len(rows),
        "headers": headers[:20],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
