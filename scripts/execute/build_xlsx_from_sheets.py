#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build a .xlsx workbook from the JSON intermediate produced by
export_wecom_sheet.js.

The JSON holds, per sheet, a sparse list of cells {r, c, v, t, f}:
  v = value (number, string, bool)
  t = cell type from the WeCom engine: 1=bool, 2=number, 3=error,
      4=string, 5=formula (we store the computed value, not the formula)
  f = original Excel number-format code (e.g. 'm"月"d"日"', '0.00%') when present

We reuse f verbatim as the openpyxl number_format so dates and percentages
render exactly as they did in the source sheet (the stored value for a date is
the Excel serial, e.g. 45663; the format code turns it back into 1月1日).

Usage:
  python scripts/execute/build_xlsx_from_sheets.py <in.sheets.json> [out.xlsx]
"""

import json
import sys
import os
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Control chars that are illegal in the XML openpyxl writes (everything below
# 0x20 except tab/newline/carriage-return, plus a couple of high control codes).
ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]'
)


def clean_text(s):
    """Strip XML-illegal control characters so openpyxl can write the cell."""
    return ILLEGAL_XML_RE.sub('', s)

# WeCom engine cell types
T_BOOL, T_NUMBER, T_ERROR, T_STRING, T_FORMULA = 1, 2, 3, 4, 5

# Excel's epoch quirk: serials >= 1 map through the 1900 date system. We only
# convert to a real date when the format code clearly asks for a date/time;
# otherwise a bare number keeps its numeric value (and its % / plain format).
DATE_HINTS = ('y', 'm', 'd', 'h', 's', '年', '月', '日', '时', '分')


def looks_like_date_format(fmt):
    if not fmt:
        return False
    f = fmt.lower()
    # A percent or pure-number format is never a date even if it contains 'm'.
    if '%' in f:
        return False
    return any(h in fmt for h in DATE_HINTS) and any(c in f for c in 'ymdhs年月日')


def flatten_rich_text(v):
    """A rich-text cell value is {'r': [{'t': '...', 'rPr': {...}}, ...]}; the
    visible text is the concatenation of each run's 't'. We keep plain text only
    (per-run bold/color isn't carried into the export)."""
    runs = v.get('r')
    if isinstance(runs, list):
        return ''.join(str(seg.get('t', '')) for seg in runs if isinstance(seg, dict))
    # Some builds nest under 'text' or expose a flat 's'/'str'.
    for k in ('text', 's', 'str', 'value'):
        if isinstance(v.get(k), str):
            return v[k]
    return ''


def coerce_value(cell):
    v = cell.get('v')
    t = cell.get('t')
    if v is None:
        return None
    if isinstance(v, dict):
        # Rich text (styled runs) — flatten to plain string regardless of type.
        return clean_text(flatten_rich_text(v))
    if isinstance(v, (list, tuple)):
        return clean_text(''.join(str(x) for x in v))
    if t == T_BOOL:
        return bool(v)
    if t == T_STRING:
        return clean_text(v) if isinstance(v, str) else v
    # number / formula-result / error: write through as-is (clean if it's a str)
    if isinstance(v, str):
        return clean_text(v)
    return v


def build(in_path, out_path):
    with open(in_path, 'r', encoding='utf-8') as fh:
        data = json.load(fh)

    wb = Workbook()
    wb.remove(wb.active)

    order = data.get('sheetOrder') or [s['name'] for s in data['sheets']]
    by_name = {s['name']: s for s in data['sheets']}

    used_titles = set()
    written = []
    for name in order:
        sheet = by_name.get(name)
        if sheet is None:
            continue
        # openpyxl forbids these chars in a sheet title and caps length at 31.
        title = name
        for ch in '[]:*?/\\':
            title = title.replace(ch, ' ')
        title = title.strip()[:31] or 'Sheet'
        base, n = title, 1
        while title in used_titles:
            n += 1
            suffix = f'~{n}'
            title = base[:31 - len(suffix)] + suffix
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        cells = sheet.get('cells', [])
        for cell in cells:
            r = cell['r'] + 1  # openpyxl is 1-based
            c = cell['c'] + 1
            value = coerce_value(cell)
            if value is None:
                continue
            try:
                xc = ws.cell(row=r, column=c, value=value)
            except Exception:
                # openpyxl rejected the value (illegal char / unexpected shape);
                # stringify + clean so content is never silently dropped.
                xc = ws.cell(row=r, column=c, value=clean_text(str(value)))
            fmt = cell.get('f')
            if fmt:
                # openpyxl wants the literal Excel code; escaped quotes from JSON
                # are already real quotes here.
                try:
                    xc.number_format = fmt
                except Exception:
                    pass

        # Apply merged ranges.
        for m in sheet.get('merges', []):
            try:
                ref = '%s%d:%s%d' % (
                    get_column_letter(m['sc'] + 1), m['sr'] + 1,
                    get_column_letter(m['ec'] + 1), m['er'] + 1,
                )
                if ':' in ref and ref.split(':')[0] != ref.split(':')[1]:
                    ws.merge_cells(ref)
            except Exception:
                pass

        written.append((title, len(cells)))

    if not wb.sheetnames:
        wb.create_sheet(title='Empty')

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return written


def main():
    if len(sys.argv) < 2:
        print('Usage: python build_xlsx_from_sheets.py <in.sheets.json> [out.xlsx]', file=sys.stderr)
        sys.exit(1)
    in_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else in_path.replace('.sheets.json', '.xlsx').replace('.json', '.xlsx')
    written = build(in_path, out_path)
    total = sum(c for _, c in written)
    print(out_path)
    print(f'sheets={len(written)} cells={total}', file=sys.stderr)
    for title, n in written:
        print(f'  {title}: {n} cells', file=sys.stderr)


if __name__ == '__main__':
    main()
